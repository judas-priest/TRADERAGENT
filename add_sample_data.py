#!/usr/bin/env python3
"""
Script to add sample data to ALMIR Backtesting Template
Adds 10 realistic test records for demonstration
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

def add_sample_data():
    """Add 10 sample test records to the template"""

    # Load the workbook
    wb = load_workbook('ALMIR_Backtesting_Template.xlsx')
    ws = wb['Raw Data']

    # Sample data (10 realistic test records)
    sample_data = [
        # Test 1: BTCUSDT 4h Default - Good performance
        ["T001", "2026-02-01", "BTCUSDT", "Large Cap", "4h", "1Y", "Default",
         14, 30, 65, 2, 30, 10, 14, 6, 10, "No", "No", 1.5, 25,
         45.5, 4550, 48, 28, 20, 58.33, None, -12.5, -568.75, 1.85, 0.95, 2.15, -1.35, 5.8, -3.2, "4h 15m",
         28, 20, None, None, 6.2, 4, "Medium", "Good performance on BTC, strong trend following"],

        # Test 2: BTCUSDT 1d Default - Excellent performance
        ["T002", "2026-02-01", "BTCUSDT", "Large Cap", "1d", "1Y", "Default",
         14, 30, 65, 2, 30, 10, 14, 6, 10, "No", "No", 1.5, 25,
         62.8, 6280, 35, 23, 12, 65.71, None, -8.3, -521.24, 2.35, 1.79, 3.25, -2.18, 8.5, -4.5, "2d 8h",
         23, 12, None, None, 6.5, 5, "Low", "Excellent on daily timeframe"],

        # Test 3: ETHUSDT 4h Default
        ["T003", "2026-02-01", "ETHUSDT", "Large Cap", "4h", "1Y", "Default",
         14, 30, 65, 2, 30, 10, 14, 6, 10, "No", "No", 1.5, 25,
         38.2, 3820, 52, 29, 23, 55.77, None, -15.2, -580.64, 1.68, 0.73, 2.05, -1.52, 6.2, -3.8, "4h 30m",
         30, 22, None, None, 5.9, 4, "Medium", "Good ETH performance"],

        # Test 4: BNBUSDT 1h Conservative
        ["T004", "2026-02-01", "BNBUSDT", "Exchange Tokens", "1h", "1Y", "Conservative",
         10, 25, 70, 5, 26, 9, 10, 8, 12, "Yes", "Yes", 2.0, 30,
         28.5, 2850, 32, 18, 14, 56.25, None, -10.8, -307.8, 1.45, 0.89, 2.42, -1.95, 5.5, -3.8, "1h 45m",
         18, 14, None, None, 7.8, 4, "Low", "Conservative approach reduces trades"],

        # Test 5: SOLUSDT 4h Balanced
        ["T005", "2026-02-02", "SOLUSDT", "Layer 1", "4h", "1Y", "Balanced",
         12, 28, 67, 3, 28, 10, 12, 6, 10, "No", "No", 1.5, 25,
         52.3, 5230, 45, 27, 18, 60.00, None, -14.5, -758.35, 1.95, 1.16, 2.85, -1.72, 7.2, -3.5, "5h 10m",
         28, 17, None, None, 6.3, 5, "Medium", "SOL shows strong momentum signals"],

        # Test 6: ADAUSDT 1d Default
        ["T006", "2026-02-02", "ADAUSDT", "Layer 1", "1d", "2Y", "Default",
         14, 30, 65, 2, 30, 10, 14, 6, 10, "No", "No", 1.5, 25,
         41.2, 4120, 38, 21, 17, 55.26, None, -18.5, -762.2, 1.52, 1.08, 3.15, -2.45, 7.8, -5.2, "2d 6h",
         22, 16, None, None, 5.8, 3, "High", "ADA has more false signals on 1d"],

        # Test 7: UNIUSDT 4h Aggressive
        ["T007", "2026-02-02", "UNIUSDT", "DeFi", "4h", "1Y", "Aggressive",
         10, 30, 65, 2, 30, 10, 14, 4, 8, "No", "No", 1.5, 25,
         48.5, 4850, 68, 38, 30, 55.88, None, -22.3, -1081.55, 1.65, 0.71, 2.28, -1.85, 6.8, -4.5, "3h 20m",
         40, 28, None, None, 4.8, 3, "High", "More trades but higher drawdown"],

        # Test 8: LINKUSDT 1h Default
        ["T008", "2026-02-02", "LINKUSDT", "DeFi", "1h", "1Y", "Default",
         14, 30, 65, 2, 30, 10, 14, 6, 10, "No", "No", 1.5, 25,
         35.8, 3580, 58, 32, 26, 55.17, None, -16.8, -601.44, 1.58, 0.62, 2.18, -1.72, 5.5, -4.2, "1h 50m",
         34, 24, None, None, 6.1, 3, "Medium", "LINK decent on 1h"],

        # Test 9: MATICUSDT 4h Default
        ["T009", "2026-02-03", "MATICUSDT", "Layer 2", "4h", "1Y", "Default",
         14, 30, 65, 2, 30, 10, 14, 6, 10, "No", "No", 1.5, 25,
         44.2, 4420, 42, 24, 18, 57.14, None, -13.2, -583.44, 1.75, 1.05, 2.65, -1.82, 6.5, -3.8, "4h 40m",
         25, 17, None, None, 6.0, 4, "Medium", "MATIC Layer 2 performing well"],

        # Test 10: DOGEUSDT 1d Conservative
        ["T010", "2026-02-03", "DOGEUSDT", "Meme Coins", "1d", "1Y", "Conservative",
         10, 25, 70, 5, 26, 9, 10, 8, 12, "Yes", "Yes", 2.0, 30,
         32.5, 3250, 22, 12, 10, 54.55, None, -11.5, -373.75, 1.42, 1.48, 3.85, -3.12, 8.2, -6.5, "3d 2h",
         13, 9, None, None, 7.5, 3, "High", "DOGE volatile, conservative helps"],
    ]

    # Write sample data to rows 2-11
    for row_idx, data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    # Add conditional formatting for Win Rate
    from openpyxl.formatting.rule import CellIsRule

    # Win Rate (Column AA): Green >60%, Yellow 50-60%, Red <50%
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.conditional_formatting.add('AA2:AA10000',
        CellIsRule(operator='greaterThan', formula=['60'], fill=green_fill))
    ws.conditional_formatting.add('AA2:AA10000',
        CellIsRule(operator='between', formula=['50', '60'], fill=yellow_fill))
    ws.conditional_formatting.add('AA2:AA10000',
        CellIsRule(operator='lessThan', formula=['50'], fill=red_fill))

    # Profit Factor (Column AB): Green >2.0, Yellow 1.5-2.0, Red <1.5
    ws.conditional_formatting.add('AB2:AB10000',
        CellIsRule(operator='greaterThan', formula=['2'], fill=green_fill))
    ws.conditional_formatting.add('AB2:AB10000',
        CellIsRule(operator='between', formula=['1.5', '2'], fill=yellow_fill))
    ws.conditional_formatting.add('AB2:AB10000',
        CellIsRule(operator='lessThan', formula=['1.5'], fill=red_fill))

    # Max Drawdown (Column Z): Green >-10%, Yellow -10% to -20%, Red <-20%
    # Note: More negative is worse
    ws.conditional_formatting.add('Z2:Z10000',
        CellIsRule(operator='greaterThan', formula=['-10'], fill=green_fill))
    ws.conditional_formatting.add('Z2:Z10000',
        CellIsRule(operator='between', formula=['-20', '-10'], fill=yellow_fill))
    ws.conditional_formatting.add('Z2:Z10000',
        CellIsRule(operator='lessThan', formula=['-20'], fill=red_fill))

    # Save the workbook
    wb.save('ALMIR_Backtesting_Template.xlsx')
    print("✅ Sample data added successfully!")
    print("   - 10 test records added to Raw Data sheet")
    print("   - Conditional formatting applied")
    print("   - Formulas will auto-calculate Win Rate and Profit Factor")

if __name__ == "__main__":
    add_sample_data()
