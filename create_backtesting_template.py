#!/usr/bin/env python3
"""
Script to create ALMIR Backtesting Results Template
Creates an Excel file with all required sheets, formulas, and formatting
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def create_backtesting_template():
    """Create the backtesting template Excel file"""

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create all required sheets
    create_raw_data_sheet(wb)
    create_summary_by_category_sheet(wb)
    create_summary_by_timeframe_sheet(wb)
    create_summary_by_pair_sheet(wb)
    create_top_performers_sheet(wb)
    create_parameter_analysis_sheet(wb)
    create_candle_analysis_sheet(wb)
    create_charts_sheet(wb)

    # Save the workbook
    wb.save('ALMIR_Backtesting_Template.xlsx')
    print("✅ ALMIR_Backtesting_Template.xlsx created successfully!")

def create_raw_data_sheet(wb):
    """Create the Raw Data sheet with all columns and validation"""

    ws = wb.create_sheet("Raw Data", 0)

    # Define headers
    headers = [
        # Identification
        "Test ID", "Date Tested", "Symbol", "Category", "Timeframe",
        "Data Period", "Parameter Preset",

        # Strategy Parameters
        "RSI Length", "RSI Oversold", "RSI Overbought",
        "MACD Fast", "MACD Slow", "MACD Signal",
        "Stochastic Length", "Min Confluence Score", "Min Bars Between Signals",
        "Use Volatility Filter", "Use Trend Filter", "ATR Multiplier", "ADX Threshold",

        # Performance Metrics
        "Net Profit (%)", "Net Profit ($)", "Total Trades", "Winning Trades", "Losing Trades",
        "Win Rate (%)", "Profit Factor", "Max Drawdown (%)", "Max Drawdown ($)",
        "Sharpe Ratio", "Average Trade (%)", "Average Win (%)", "Average Loss (%)",
        "Largest Win (%)", "Largest Loss (%)", "Average Trade Duration",

        # Additional Statistics
        "Green Candle Trades", "Red Candle Trades", "Green Candle Win Rate (%)",
        "Red Candle Win Rate (%)", "Avg Confluence Score",

        # Qualitative Observations
        "Visual Quality (1-5)", "False Signals", "Notes"
    ]

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Set column widths
    column_widths = {
        'A': 10, 'B': 12, 'C': 12, 'D': 14, 'E': 12, 'F': 12, 'G': 16,
        'H': 11, 'I': 13, 'J': 15, 'K': 11, 'L': 11, 'M': 12, 'N': 16,
        'O': 18, 'P': 22, 'Q': 18, 'R': 16, 'S': 14, 'T': 14,
        'U': 14, 'V': 14, 'W': 13, 'X': 13, 'Y': 13, 'Z': 12,
        'AA': 14, 'AB': 14, 'AC': 14, 'AD': 16, 'AE': 16, 'AF': 16,
        'AG': 16, 'AH': 16, 'AI': 18, 'AJ': 18, 'AK': 20,
        'AL': 22, 'AM': 22, 'AN': 18, 'AO': 14, 'AP': 30
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze first row
    ws.freeze_panes = "A2"

    # Add data validation for dropdowns
    add_data_validations(ws)

    # Add formulas for calculated columns
    add_formulas_to_raw_data(ws)

def add_data_validations(ws):
    """Add data validation dropdowns to the Raw Data sheet"""

    # Category dropdown
    category_dv = DataValidation(type="list", formula1='"Large Cap,DeFi,Layer 1,Layer 2,Meme Coins,Gaming/Metaverse,Exchange Tokens"', allow_blank=True)
    ws.add_data_validation(category_dv)
    category_dv.add('D2:D10000')

    # Timeframe dropdown
    timeframe_dv = DataValidation(type="list", formula1='"1m,5m,15m,30m,1h,4h,1d,1w"', allow_blank=True)
    ws.add_data_validation(timeframe_dv)
    timeframe_dv.add('E2:E10000')

    # Data Period dropdown
    period_dv = DataValidation(type="list", formula1='"6M,1Y,2Y,3Y"', allow_blank=True)
    ws.add_data_validation(period_dv)
    period_dv.add('F2:F10000')

    # Parameter Preset dropdown
    preset_dv = DataValidation(type="list", formula1='"Default,Conservative,Balanced,Aggressive,Custom"', allow_blank=True)
    ws.add_data_validation(preset_dv)
    preset_dv.add('G2:G10000')

    # Visual Quality dropdown
    quality_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(quality_dv)
    quality_dv.add('AN2:AN10000')

    # False Signals dropdown
    signals_dv = DataValidation(type="list", formula1='"Low,Medium,High"', allow_blank=True)
    ws.add_data_validation(signals_dv)
    signals_dv.add('AO2:AO10000')

    # Boolean dropdowns for filters
    bool_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(bool_dv)
    bool_dv.add('Q2:Q10000')  # Use Volatility Filter

    bool_dv2 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(bool_dv2)
    bool_dv2.add('R2:R10000')  # Use Trend Filter

def add_formulas_to_raw_data(ws):
    """Add formulas for calculated columns in Raw Data sheet"""

    # Starting from row 2 (after header)
    for row in range(2, 102):  # Add formulas for 100 rows
        # Win Rate (%) - Column AA (column 27)
        # Formula: =IF(AND(W{row}<>"", W{row}>0), X{row}/W{row}*100, "")
        ws[f'AA{row}'] = f'=IF(AND(W{row}<>"", W{row}>0), X{row}/W{row}*100, "")'

        # Profit Factor - Column AB (column 28)
        # Formula: =IF(AND(X{row}>0, Y{row}>0, AC{row}<>"", AD{row}<>""), (X{row}*AC{row})/(Y{row}*ABS(AD{row})), "")
        ws[f'AB{row}'] = f'=IF(AND(X{row}>0, Y{row}>0, AC{row}<>"", AD{row}<>""), (X{row}*AC{row})/(Y{row}*ABS(AD{row})), "")'

        # Green Candle Win Rate (%) - Column AL
        # Formula: =IF(AND(AJ{row}<>"", AJ{row}>0), (green candle wins calculation), "")
        # This would need tracking wins per candle type - simplified for now
        ws[f'AL{row}'] = f'=IF(AJ{row}<>"", "", "")'

        # Red Candle Win Rate (%) - Column AM
        ws[f'AM{row}'] = f'=IF(AK{row}<>"", "", "")'

def create_summary_by_category_sheet(wb):
    """Create Summary by Category sheet"""

    ws = wb.create_sheet("Summary by Category")

    # Headers
    headers = ["Category", "Avg Win Rate (%)", "Avg Profit Factor", "Avg Net Profit (%)",
               "Best Timeframe", "Count of Tests", "Top 3 Pairs"]

    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Categories
    categories = ["Large Cap", "DeFi", "Layer 1", "Layer 2", "Meme Coins", "Gaming/Metaverse", "Exchange Tokens"]

    for row_num, category in enumerate(categories, 2):
        ws.cell(row=row_num, column=1).value = category

        # Add AVERAGEIF formulas
        # Avg Win Rate
        ws.cell(row=row_num, column=2).value = f'=AVERAGEIF(\'Raw Data\'!$D:$D, A{row_num}, \'Raw Data\'!$AA:$AA)'

        # Avg Profit Factor
        ws.cell(row=row_num, column=3).value = f'=AVERAGEIF(\'Raw Data\'!$D:$D, A{row_num}, \'Raw Data\'!$AB:$AB)'

        # Avg Net Profit
        ws.cell(row=row_num, column=4).value = f'=AVERAGEIF(\'Raw Data\'!$D:$D, A{row_num}, \'Raw Data\'!$U:$U)'

        # Count of Tests
        ws.cell(row=row_num, column=6).value = f'=COUNTIF(\'Raw Data\'!$D:$D, A{row_num})'

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 30

    ws.freeze_panes = "A2"

def create_summary_by_timeframe_sheet(wb):
    """Create Summary by Timeframe sheet"""

    ws = wb.create_sheet("Summary by Timeframe")

    # Headers
    headers = ["Timeframe", "Avg Win Rate (%)", "Avg Profit Factor", "Avg Net Profit (%)",
               "Best Category", "Count of Tests", "Avg Trade Duration"]

    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Timeframes
    timeframes = ["1m", "5m", "15m", "30m", "1h", "4h", "1d", "1w"]

    for row_num, tf in enumerate(timeframes, 2):
        ws.cell(row=row_num, column=1).value = tf

        # Add AVERAGEIF formulas
        ws.cell(row=row_num, column=2).value = f'=AVERAGEIF(\'Raw Data\'!$E:$E, A{row_num}, \'Raw Data\'!$AA:$AA)'
        ws.cell(row=row_num, column=3).value = f'=AVERAGEIF(\'Raw Data\'!$E:$E, A{row_num}, \'Raw Data\'!$AB:$AB)'
        ws.cell(row=row_num, column=4).value = f'=AVERAGEIF(\'Raw Data\'!$E:$E, A{row_num}, \'Raw Data\'!$U:$U)'
        ws.cell(row=row_num, column=6).value = f'=COUNTIF(\'Raw Data\'!$E:$E, A{row_num})'

    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18

    ws.freeze_panes = "A2"

def create_summary_by_pair_sheet(wb):
    """Create Summary by Pair sheet"""

    ws = wb.create_sheet("Summary by Pair")

    # Headers
    headers = ["Symbol", "Best Win Rate (%)", "Best Profit Factor", "Best Timeframe",
               "Best Preset", "Count of Tests", "Avg Performance"]

    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_font = Font(bold=True, size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Instructions for user
    ws.cell(row=2, column=1).value = "Unique symbols from Raw Data will appear here"
    ws.cell(row=2, column=1).font = Font(italic=True, color="808080")

    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18

    ws.freeze_panes = "A2"

def create_top_performers_sheet(wb):
    """Create Top Performers sheet"""

    ws = wb.create_sheet("Top Performers")

    # Section 1: Top 20 by Win Rate
    ws.cell(row=1, column=1).value = "Top 20 by Win Rate"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:H1')

    headers = ["Rank", "Symbol", "Timeframe", "Preset", "Win Rate (%)", "Profit Factor", "Net Profit (%)", "Max DD (%)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Section 2: Top 20 by Profit Factor (starting at row 25)
    ws.cell(row=25, column=1).value = "Top 20 by Profit Factor"
    ws.cell(row=25, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=25, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells('A25:H25')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=26, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Section 3: Top 20 by Net Profit (starting at row 49)
    ws.cell(row=49, column=1).value = "Top 20 by Net Profit (%)"
    ws.cell(row=49, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=49, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws.merge_cells('A49:H49')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=50, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15

def create_parameter_analysis_sheet(wb):
    """Create Parameter Analysis sheet"""

    ws = wb.create_sheet("Parameter Analysis")

    ws.cell(row=1, column=1).value = "Parameter Analysis & Correlation"
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    # Analysis sections
    ws.cell(row=3, column=1).value = "Confluence Score Analysis"
    ws.cell(row=3, column=1).font = Font(bold=True, size=12)

    headers = ["Min Confluence Score", "Avg Win Rate (%)", "Avg Profit Factor", "Count of Tests"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Confluence scores 4-8
    for i, score in enumerate([4, 5, 6, 7, 8], 5):
        ws.cell(row=i, column=1).value = score
        ws.cell(row=i, column=2).value = f'=AVERAGEIF(\'Raw Data\'!$O:$O, A{i}, \'Raw Data\'!$AA:$AA)'
        ws.cell(row=i, column=3).value = f'=AVERAGEIF(\'Raw Data\'!$O:$O, A{i}, \'Raw Data\'!$AB:$AB)'
        ws.cell(row=i, column=4).value = f'=COUNTIF(\'Raw Data\'!$O:$O, A{i})'

    # RSI Length Analysis
    ws.cell(row=11, column=1).value = "RSI Length Analysis"
    ws.cell(row=11, column=1).font = Font(bold=True, size=12)

    for col_num, header in enumerate(["RSI Length", "Avg Win Rate (%)", "Avg Profit Factor", "Count of Tests"], 1):
        cell = ws.cell(row=12, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Set column widths
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20

def create_candle_analysis_sheet(wb):
    """Create Candle Analysis sheet"""

    ws = wb.create_sheet("Candle Analysis")

    ws.cell(row=1, column=1).value = "Green vs Red Candle Performance Analysis"
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.merge_cells('A1:G1')

    # Overall comparison
    ws.cell(row=3, column=1).value = "Overall Performance"
    ws.cell(row=3, column=1).font = Font(bold=True, size=12)

    headers = ["Candle Type", "Total Trades", "Avg Win Rate (%)", "Avg Profit Factor", "Notes"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Green and Red rows
    ws.cell(row=5, column=1).value = "Green Candles (Long)"
    ws.cell(row=5, column=2).value = '=SUM(\'Raw Data\'!$AJ:$AJ)'
    ws.cell(row=5, column=3).value = '=AVERAGE(\'Raw Data\'!$AL:$AL)'

    ws.cell(row=6, column=1).value = "Red Candles (Short)"
    ws.cell(row=6, column=2).value = '=SUM(\'Raw Data\'!$AK:$AK)'
    ws.cell(row=6, column=3).value = '=AVERAGE(\'Raw Data\'!$AM:$AM)'

    # By Category analysis
    ws.cell(row=9, column=1).value = "Performance by Category"
    ws.cell(row=9, column=1).font = Font(bold=True, size=12)

    headers = ["Category", "Green Candle WR (%)", "Red Candle WR (%)", "Difference"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 22

def create_charts_sheet(wb):
    """Create Charts & Visualizations sheet"""

    ws = wb.create_sheet("Charts & Visualizations")

    ws.cell(row=1, column=1).value = "Charts & Visualizations"
    ws.cell(row=1, column=1).font = Font(bold=True, size=18)
    ws.merge_cells('A1:F1')

    # Instructions
    ws.cell(row=3, column=1).value = "📊 Recommended Charts to Create:"
    ws.cell(row=3, column=1).font = Font(bold=True, size=14)

    instructions = [
        "",
        "1. Win Rate by Category (Bar Chart)",
        "   - Data: Summary by Category sheet, columns A & B",
        "   - Chart Type: Clustered Column",
        "",
        "2. Profit Factor by Timeframe (Line Chart)",
        "   - Data: Summary by Timeframe sheet, columns A & C",
        "   - Chart Type: Line with Markers",
        "",
        "3. Net Profit vs Max Drawdown (Scatter Plot)",
        "   - Data: Raw Data sheet, columns U & Z",
        "   - Chart Type: Scatter",
        "   - Color by: Category (column D)",
        "",
        "4. Category vs Timeframe Heatmap",
        "   - Create pivot table with Category (rows) and Timeframe (columns)",
        "   - Values: Average Win Rate",
        "   - Apply conditional formatting (color scale)",
        "",
        "5. Confluence Score Distribution (Histogram)",
        "   - Data: Parameter Analysis sheet",
        "   - Chart Type: Column Chart",
        "",
        "6. Green vs Red Candle Performance (Comparison Chart)",
        "   - Data: Candle Analysis sheet",
        "   - Chart Type: Clustered Bar",
        "",
        "📌 Note: Charts can be created manually in Excel/Google Sheets",
        "   using Insert > Chart and selecting the recommended data ranges."
    ]

    for i, instruction in enumerate(instructions, 4):
        ws.cell(row=i, column=1).value = instruction
        if instruction.startswith(("1.", "2.", "3.", "4.", "5.", "6.")):
            ws.cell(row=i, column=1).font = Font(bold=True, size=11)

    # Set column width
    ws.column_dimensions['A'].width = 80

if __name__ == "__main__":
    create_backtesting_template()
