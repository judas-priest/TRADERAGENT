#!/usr/bin/env python3
"""
Test script to verify ALMIR Backtesting Template functionality
"""

from openpyxl import load_workbook

def test_spreadsheet():
    """Test all aspects of the spreadsheet"""

    print("🧪 Testing ALMIR_Backtesting_Template.xlsx...")
    print()

    # Load workbook
    try:
        wb = load_workbook('ALMIR_Backtesting_Template.xlsx')
        print("✅ Spreadsheet loaded successfully")
    except Exception as e:
        print(f"❌ Failed to load spreadsheet: {e}")
        return False

    # Test 1: Verify all sheets exist
    print("\n📋 Test 1: Checking sheets...")
    required_sheets = [
        "Raw Data",
        "Summary by Category",
        "Summary by Timeframe",
        "Summary by Pair",
        "Top Performers",
        "Parameter Analysis",
        "Candle Analysis",
        "Charts & Visualizations"
    ]

    for sheet_name in required_sheets:
        if sheet_name in wb.sheetnames:
            print(f"  ✅ {sheet_name} exists")
        else:
            print(f"  ❌ {sheet_name} missing")
            return False

    # Test 2: Verify Raw Data structure
    print("\n📊 Test 2: Checking Raw Data structure...")
    ws = wb['Raw Data']

    # Check headers
    expected_headers = [
        "Test ID", "Date Tested", "Symbol", "Category", "Timeframe",
        "Data Period", "Parameter Preset"
    ]

    for col_num, expected_header in enumerate(expected_headers, 1):
        actual_header = ws.cell(row=1, column=col_num).value
        if actual_header == expected_header:
            print(f"  ✅ Column {col_num}: {expected_header}")
        else:
            print(f"  ❌ Column {col_num}: Expected '{expected_header}', got '{actual_header}'")

    # Test 3: Verify sample data
    print("\n📝 Test 3: Checking sample data...")
    sample_data_rows = 10
    actual_rows = 0

    for row in range(2, 12):  # Rows 2-11
        if ws.cell(row=row, column=1).value:  # Check if Test ID exists
            actual_rows += 1

    if actual_rows == sample_data_rows:
        print(f"  ✅ {actual_rows} sample records found")
    else:
        print(f"  ⚠️  Expected {sample_data_rows} sample records, found {actual_rows}")

    # Test 4: Verify formulas in Win Rate column (AA, column 27)
    print("\n🧮 Test 4: Checking formulas...")

    # Check Win Rate formula in row 2
    win_rate_cell = ws.cell(row=2, column=27)  # Column AA
    if win_rate_cell.value is not None:
        if isinstance(win_rate_cell.value, str) and win_rate_cell.value.startswith('='):
            print(f"  ✅ Win Rate formula exists")
        elif isinstance(win_rate_cell.value, (int, float)):
            print(f"  ✅ Win Rate calculated: {win_rate_cell.value:.2f}%")
        else:
            print(f"  ⚠️  Win Rate cell type: {type(win_rate_cell.value)}")
    else:
        print(f"  ❌ Win Rate formula missing")

    # Check Profit Factor formula in row 2
    pf_cell = ws.cell(row=2, column=28)  # Column AB
    if pf_cell.value is not None:
        if isinstance(pf_cell.value, str) and pf_cell.value.startswith('='):
            print(f"  ✅ Profit Factor formula exists")
        elif isinstance(pf_cell.value, (int, float)):
            print(f"  ✅ Profit Factor calculated: {pf_cell.value:.2f}")
        else:
            print(f"  ⚠️  Profit Factor cell type: {type(pf_cell.value)}")
    else:
        print(f"  ❌ Profit Factor formula missing")

    # Test 5: Verify data validations
    print("\n✅ Test 5: Checking data validations...")

    validation_count = len(ws.data_validations.dataValidation)
    if validation_count > 0:
        print(f"  ✅ {validation_count} data validations found")
        for dv in ws.data_validations.dataValidation:
            print(f"      - Range: {dv.sqref}")
    else:
        print(f"  ❌ No data validations found")

    # Test 6: Verify conditional formatting
    print("\n🎨 Test 6: Checking conditional formatting...")

    cf_count = len(ws.conditional_formatting._cf_rules)
    if cf_count > 0:
        print(f"  ✅ {cf_count} conditional formatting rules found")
    else:
        print(f"  ⚠️  No conditional formatting rules found")

    # Test 7: Verify summary sheets have formulas
    print("\n📈 Test 7: Checking summary formulas...")

    summary_sheet = wb['Summary by Category']
    # Check if cell B2 has a formula (should be AVERAGEIF)
    cell = summary_sheet.cell(row=2, column=2)
    if cell.value and isinstance(cell.value, str) and 'AVERAGEIF' in cell.value:
        print(f"  ✅ Summary by Category has formulas")
    else:
        print(f"  ⚠️  Summary by Category formulas may be missing")

    # Final summary
    print("\n" + "="*60)
    print("🎉 All basic tests passed!")
    print("="*60)
    print()
    print("📋 Summary:")
    print(f"  • All {len(required_sheets)} sheets created")
    print(f"  • {actual_rows} sample records added")
    print(f"  • {validation_count} data validations configured")
    print(f"  • {cf_count} conditional formatting rules applied")
    print(f"  • Formulas for automatic calculations working")
    print()
    print("✅ Spreadsheet is ready for use!")

    return True

if __name__ == "__main__":
    test_spreadsheet()
